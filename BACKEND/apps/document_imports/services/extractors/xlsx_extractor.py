from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .base_extractor import (
    BaseExtractor,
    ExtractedPage,
    ExtractionResult,
)


class XLSXExtractor(BaseExtractor):
    """
    Extracteur de fichiers Excel (.xlsx).

    Extrait :
    - les feuilles ;
    - les en-têtes ;
    - les lignes ;
    - les données brutes ;
    - quelques métadonnées.
    """

    SUPPORTED_EXTENSIONS = {
        ".xlsx",
    }

    def supports(
        self,
        file_path: str,
    ) -> bool:
        """
        Vérifie si le fichier est un XLSX.
        """

        extension = (
            Path(file_path)
            .suffix
            .lower()
        )

        return extension in self.SUPPORTED_EXTENSIONS

    def extract(
        self,
        file_path: str,
    ) -> ExtractionResult:
        """
        Extrait le contenu du fichier Excel.
        """

        if not self.supports(file_path):
            return ExtractionResult(
                success=False,
                error="Le fichier fourni n'est pas un fichier XLSX.",
            )

        workbook = None

        try:
            workbook = load_workbook(
                filename=file_path,
                data_only=True,
            )

            pages = []

            page_number = 0

            for worksheet in workbook.worksheets:

                page_number += 1

                tables = self._extract_worksheet(
                    worksheet
                )

                text = self._worksheet_to_text(
                    worksheet
                )

                page = ExtractedPage(
                    page_number=page_number,
                    text=text,
                    tables=tables,
                    metadata={
                        "source_type": "xlsx",
                        "sheet_name": worksheet.title,
                        "max_row": worksheet.max_row,
                        "max_column": worksheet.max_column,
                    },
                )

                pages.append(page)

            metadata = {
                "sheet_count": len(
                    workbook.worksheets
                ),
                "sheet_names": [
                    worksheet.title
                    for worksheet in workbook.worksheets
                ],
            }

            return ExtractionResult(
                success=True,
                pages=pages,
                page_count=len(pages),
                metadata=metadata,
            )

        except Exception as exc:

            return ExtractionResult(
                success=False,
                error=str(exc),
            )

        finally:

            if workbook is not None:
                workbook.close()

    def _extract_worksheet(
        self,
        worksheet: Any,
    ) -> list[dict[str, Any]]:
        """
        Transforme une feuille Excel en structure
        compatible avec ExtractedTable.
        """

        rows = []

        for row in worksheet.iter_rows(
            values_only=True
        ):

            normalized_row = [
                self._normalize_cell(value)
                for value in row
            ]

            # On ignore les lignes complètement vides.
            if any(
                value != ""
                for value in normalized_row
            ):
                rows.append(normalized_row)

        if not rows:
            return []

        headers = rows[0]

        data_rows = rows[1:]

        return [
            {
                "table_index": 1,
                "headers": headers,
                "rows": data_rows,
                "raw_data": {
                    "sheet_name": worksheet.title,
                    "rows": rows,
                },
            }
        ]

    def _worksheet_to_text(
        self,
        worksheet: Any,
    ) -> str:
        """
        Convertit le contenu de la feuille en texte.

        Cela permettra ensuite à notre moteur d'analyse
        de traiter les données textuelles de la même manière
        que celles provenant d'un PDF ou d'un DOCX.
        """

        lines = []

        for row in worksheet.iter_rows(
            values_only=True
        ):

            values = [
                self._normalize_cell(value)
                for value in row
            ]

            if any(
                value != ""
                for value in values
            ):
                lines.append(
                    " | ".join(values)
                )

        return "\n".join(lines)

    def _normalize_cell(
        self,
        value: Any,
    ) -> str:
        """
        Convertit proprement une cellule Excel
        en texte.
        """

        if value is None:
            return ""

        if hasattr(value, "isoformat"):
            try:
                return value.isoformat()
            except Exception:
                pass

        return str(value).strip()